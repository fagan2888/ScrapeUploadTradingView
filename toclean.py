# # 1. Import all library

from requests_html import HTMLSession
import datetime
import time
import xlwt 
from xlwt import Workbook 
from datetime import datetime
from datetime import datetime
datestring = datetime.strftime(datetime.now(), '%Y;%m;%d')
import xlwings as xw
from openpyxl import Workbook, load_workbook
from selenium import webdriver
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import autoit
import xlsxwriter
from openpyxl.formula.translate import Translator
from dateutil.parser import parse
import pandas as pd
from selenium.webdriver.common.action_chains import ActionChains
import autopy
print('library')
print('=======1. Library has been imported')

# 2. get data from website (not string yet)
url = 'https://www.tradingview.com/symbols/MYX-FCPO1%21/'
session = HTMLSession()
r = session.get(url)

r.html.render()
time.sleep(20)

trd_time = 'span.js-symbol-lp-time' 
num_curr = 'div.tv-symbol-price-quote__value.js-symbol-last'
num_prev = 'div.tv-fundamental-block__value.js-symbol-prev-close'
num_open = 'div.tv-fundamental-block__value.js-symbol-open'

print('Value')
print('=======2.all value as been added (not converted to string)')

# 3. Convert all get value to string
result_server_time = r.html.find(trd_time, first=True).text
result_curr = r.html.find(num_curr, first=True).text
result_prev = r.html.find(num_prev, first=True).text
result_open = r.html.find(num_open, first=True).text

print('converted string not shown')
print('=======3.all value as been converted to string')
# 4. Print all strings

print(result_server_time)
print(result_curr)
print(result_prev)
print(result_open)

print('=======4.All String Has Been printed')



# 5. pilih di date hanya string index nomor 1 s.d. 13 -> supaya jadi : 23 Jan 15:22

index_1_to_7 = result_server_time[1:7]


print('index 1 to 13')
print('=======5. set index ')


# 6. tambahkan tahun dan detik di date

tahun = '2020 '
detik = ':00'

add_tahun = str(tahun + index_1_to_7)
print('add_tahun')
print("=======6. print clean server time" )

# 7. Print semuanya lagi buat ngecheck
print('===============')
print(add_tahun)
print(result_curr)
print(result_prev)
print(result_open)

print('===========7 . print all again')
# 8. sudah dapat stringnya dari website,
#    lalu CONVERTLAH STRING MENJADI datetime

import datetime
date_time_str = add_tahun
date_time_obj = datetime.datetime.strptime(date_time_str, '%Y %b %d')

print(type(date_time_obj))
print(date_time_obj)
print('===================8. print data type of time')
# print('Date:', date_time_obj.date())
# print('Time:', date_time_obj.time())
# print('Date-time:', date_time_obj)



# 9. sudah terconvert, lalu masuk ke excel.
#---------namain excel---
#  ambil date and month only:

judul = result_server_time[1:8]
print(judul)
judul_time = datetime.datetime.strptime(judul,'%b %d ')
date_and_month = judul_time.strftime('%m-%d')

 
out_filename = 'macro-uploader2020-'+date_and_month+".xlsx"
print(out_filename)
print('========9. done')

# 10. kita load original workbook sebagai format dan write
wb = load_workbook('original.xlsx')
ws = wb['Sheet 1']
ws['C2'] = date_time_obj
ws['D2'] = int(result_curr)
wb.save(out_filename)
print('========10. done')

# 11. Upload to website

print('doing step 11, upload to website')
DRIVER = 'chromedriver'
options = webdriver.ChromeOptions()
if os.name == "nt":
  
    options.add_argument("--start-maximized")
else:
    
    options.add_argument("--kiosk")

options.add_argument("user-data-dir=C:/Users/Rifan/AppData/Local/Google/Chrome/User Data/Default/") # Path to your chrome profile or you can open chrome and type: "chrome://version/" on URL

driver = webdriver.Chrome(DRIVER, options = options)
driver.get('https://pasardana.id/admin/macro/data')		
time.sleep(5)
wait = WebDriverWait(driver, 10)
wait.until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Upload"]'))).click()
time.sleep(3)
driver.find_element_by_css_selector("input[ng-model='files']").send_keys("D:/Project Kantor/ScrapeUploadTradingView/macro-uploader2020-01-24.xlsx") 

print('file has been imported');

#12. klik button save
autopy.mouse.smooth_move(0,0)
autopy.mouse.smooth_move(580,495)
autopy.mouse.click()

print('button save has been clicked')
# wait2 = WebDriverWait(driver, 10)
# wait2.until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Save"]')))
# ActionChains(driver).move_to_element(wait2).submit(wait2).perform()


# driver.implicitly_wait(10)
# ActionChains(driver).move_to_element(button_save).click(button).perform()
# driver.find_element_by_css_selector("button[ng-click='CRUDCtrl.toggleFormUpload(true)']").click()
# element_save.click()
time.sleep(5)
driver.close()